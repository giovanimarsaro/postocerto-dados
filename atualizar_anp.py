"""
Robo que busca o levantamento semanal de precos de combustiveis da ANP,
pega a planilha mais recente disponibilizada no site oficial, e extrai a
media nacional de Gasolina, Etanol e Diesel para o arquivo anp-referencia.json.

Se algo falhar (ANP fora do ar, mudanca de layout, etc), o script NAO apaga
o arquivo antigo - assim o app sempre tem um numero de referencia, mesmo
que nao seja o mais novo daquela semana.
"""
import json
import re
import sys
from datetime import datetime, timezone

import requests

PAGINA_ANP = "https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/precos/levantamento-de-precos-de-combustiveis-ultimas-semanas-pesquisadas"
SAIDA = "anp-referencia.json"

# nomes de produto como costumam aparecer nas planilhas da ANP
ALIAS_PRODUTOS = {
    "gasolina": ["GASOLINA COMUM", "GASOLINA C COMUM", "GASOLINA"],
    "etanol": ["ETANOL HIDRATADO", "ETANOL"],
    "diesel": ["OLEO DIESEL S10", "ÓLEO DIESEL S10", "OLEO DIESEL", "ÓLEO DIESEL"],
}


def achar_link_planilha_mais_recente():
    resp = requests.get(PAGINA_ANP, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
    resp.raise_for_status()
    # procura o primeiro link "resumo_semanal_lpc...xlsx" (a pagina lista da semana mais nova pra mais antiga)
    matches = re.findall(r'href="([^"]*resumo[_-]semanal[_-]lpc[^"]*\.xlsx)"', resp.text, re.IGNORECASE)
    if not matches:
        raise RuntimeError("Nao encontrei nenhum link de planilha na pagina da ANP. O layout pode ter mudado.")
    link = matches[0]
    if link.startswith("http"):
        return link
    return "https://www.gov.br" + link


def baixar_planilha(url):
    resp = requests.get(url, timeout=60, headers={"User-Agent": "Mozilla/5.0"})
    resp.raise_for_status()
    nome_arquivo = "planilha_anp.xlsx"
    with open(nome_arquivo, "wb") as f:
        f.write(resp.content)
    return nome_arquivo


def extrair_medias_nacionais(caminho_planilha):
    import openpyxl

    wb = openpyxl.load_workbook(caminho_planilha, data_only=True)
    resultado = {}

    for aba in wb.sheetnames:
        ws = wb[aba]
        linhas = list(ws.iter_rows(values_only=True))
        if not linhas:
            continue
        cabecalho = [str(c).strip().upper() if c else "" for c in linhas[0]]

        col_produto = next((i for i, c in enumerate(cabecalho) if "PRODUTO" in c), None)
        col_regiao = next((i for i, c in enumerate(cabecalho) if c in ("ESTADO", "REGIAO", "REGIÃO", "MUNICIPIO", "MUNICÍPIO", "ABRANGENCIA", "ABRANGÊNCIA")), None)
        col_preco = next((i for i, c in enumerate(cabecalho) if "PRECO MEDIO REVENDA" in c or "PREÇO MÉDIO REVENDA" in c or "PRECO MEDIO" in c or "PREÇO MÉDIO" in c), None)

        if col_produto is None or col_preco is None:
            continue

        for linha in linhas[1:]:
            if len(linha) <= max(col_produto, col_preco):
                continue
            produto = str(linha[col_produto]).strip().upper() if linha[col_produto] else ""
            # se a planilha tiver coluna de regiao, so aceitamos a linha "BRASIL" (media nacional)
            if col_regiao is not None:
                regiao = str(linha[col_regiao]).strip().upper() if linha[col_regiao] else ""
                if regiao and regiao != "BRASIL":
                    continue
            preco = linha[col_preco]
            if not isinstance(preco, (int, float)):
                continue
            for chave, aliases in ALIAS_PRODUTOS.items():
                if produto in aliases and chave not in resultado:
                    resultado[chave] = round(float(preco), 3)

        if len(resultado) == 3:
            break

    faltando = [k for k in ALIAS_PRODUTOS if k not in resultado]
    if faltando:
        raise RuntimeError(f"Nao consegui achar preco nacional para: {faltando}. Pode ser preciso ajustar o script pro layout atual da planilha.")

    return resultado


def carregar_json_existente():
    try:
        with open(SAIDA, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return None


def main():
    try:
        link = achar_link_planilha_mais_recente()
        print(f"Planilha encontrada: {link}")
        caminho = baixar_planilha(link)
        medias = extrair_medias_nacionais(caminho)

        dados = {
            "gasolina": medias["gasolina"],
            "etanol": medias["etanol"],
            "diesel": medias["diesel"],
            "fonte": link,
            "atualizado_em": datetime.now(timezone.utc).isoformat(),
        }
        with open(SAIDA, "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=2)
        print("anp-referencia.json atualizado com sucesso:", dados)

    except Exception as e:
        print(f"ERRO ao atualizar dados da ANP: {e}", file=sys.stderr)
        anterior = carregar_json_existente()
        if anterior:
            print("Mantendo o arquivo anterior sem alteracoes.")
        else:
            # primeira execucao e falhou: cria um arquivo com valores de referencia neutros
            dados = {
                "gasolina": 6.09, "etanol": 4.42, "diesel": 6.31,
                "fonte": "valor inicial (robo ainda nao conseguiu buscar da ANP)",
                "atualizado_em": datetime.now(timezone.utc).isoformat(),
            }
            with open(SAIDA, "w", encoding="utf-8") as f:
                json.dump(dados, f, ensure_ascii=False, indent=2)
        sys.exit(1)


if __name__ == "__main__":
    main()
